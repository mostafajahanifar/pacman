import os

import matplotlib.colors as mcolors
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
from openpyxl.styles import PatternFill

from pacman.config import DATA_DIR, RESULTS_DIR

# Directory containing the cancer types' subdirectories
base_dir = os.path.join(RESULTS_DIR, "genomic/cnv_anova")

# Create a writer to write results into an Excel file
output_file = os.path.join(
    RESULTS_DIR, "genomic/cnv_anova/gene_cnv_anova_aggregated.xlsx"
)
writer = pd.ExcelWriter(output_file, engine="openpyxl")


# Function to get color based on F-statistics
def get_color_from_fstat(fstat_value, cmap, norm):
    rgba = cmap(norm(fstat_value))
    # Convert the RGBA color to hexadecimal format recognized by openpyxl
    r, g, b, _ = [int(255 * v) for v in rgba]  # Ignore alpha value
    return f"{r:02x}{g:02x}{b:02x}"


# Function to apply formatting to p-value table based on F-statistics
def format_pval_table(fstat_matrix, pval_matrix, sheet, cmap, norm):
    # Write the gene names to the first column and format the p-value table
    for row_idx, gene in enumerate(
        pval_matrix.index, start=2
    ):  # start=2 because row 1 is header
        sheet.cell(row=row_idx, column=1).value = (
            gene  # Write gene name in the first column
        )
        for col_idx, feature in enumerate(
            pval_matrix.columns, start=2
        ):  # start=2 because col 1 is index
            pval_value = pval_matrix.at[gene, feature]
            fstat_value = fstat_matrix.at[gene, feature]

            # Get the color based on the F-statistics value
            color_hex = get_color_from_fstat(fstat_value, cmap, norm)
            fill = PatternFill(
                start_color=color_hex, end_color=color_hex, fill_type="solid"
            )

            # Write the p-value (already -log transformed) into the Excel cell
            cell = sheet.cell(row=row_idx, column=col_idx)
            cell.value = f"{pval_value:.2f}"  # Format the p-value to 2 decimals
            cell.fill = fill  # Apply the color to the cell


# Colormap for F-statistics ("YlGn" colormap) with vmin=1 and vmax=50
cmap = plt.get_cmap("YlGn")
norm = mcolors.Normalize(vmin=1, vmax=50)


# Function to process each cancer type for ANOVA results
def process_anova_cancer_type(cancer_type, fstat_file, pval_file, writer, cmap, norm):
    # Read F-statistics and p-value matrices
    fstat_matrix = pd.read_csv(fstat_file, index_col=0)
    pval_matrix = pd.read_csv(pval_file, index_col=0)

    # Transpose to have genes as index and features as columns
    fstat_matrix = fstat_matrix.T
    pval_matrix = pval_matrix.T

    # Filter genes with at least one p-value < 0.01
    significant_genes = pval_matrix[pval_matrix < 0.01].dropna(how="all").index
    fstat_matrix = fstat_matrix.loc[significant_genes]
    pval_matrix = pval_matrix.loc[significant_genes]

    # Convert p-values to -log(p-values)
    pval_matrix = -np.log10(pval_matrix)

    # Sort genes based on the maximum F-statistics across features
    max_fstat_values = fstat_matrix.max(axis=1)
    sorted_genes = max_fstat_values.sort_values(ascending=False).index

    # Reorder the matrices based on sorted genes
    fstat_matrix = fstat_matrix.loc[sorted_genes]
    pval_matrix = pval_matrix.loc[sorted_genes]

    # Write the p-value matrix to an Excel sheet with color formatting based on F-statistics
    workbook = writer.book
    sheet = workbook.create_sheet(cancer_type)

    # Write the headers (features)
    sheet.append(["Gene"] + list(pval_matrix.columns))

    # Format and write the p-value table with color formatting
    format_pval_table(fstat_matrix, pval_matrix, sheet, cmap, norm)


# Loop through each cancer type's directory
for cancer_type in sorted(os.listdir(base_dir)):
    cancer_dir = os.path.join(base_dir, cancer_type)

    if os.path.isdir(cancer_dir):
        fstat_file = os.path.join(cancer_dir, f"cnv_annova-f_{cancer_type}.csv")
        pval_file = os.path.join(cancer_dir, f"cnv_annova-p_{cancer_type}.csv")

        if os.path.exists(fstat_file) and os.path.exists(pval_file):
            process_anova_cancer_type(
                cancer_type, fstat_file, pval_file, writer, cmap, norm
            )

# Save the Excel file
writer.close()

print(f"ANOVA results aggregated and formatted p-value table saved to {output_file}")
